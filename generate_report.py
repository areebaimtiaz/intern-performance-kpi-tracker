"""
generate_report.py
-------------------
Builds the monthly KPI report as a formatted Excel workbook.
Reads from the 'interns_cleaned' table inside interns.db.

Usage:
    python generate_report.py "June 2026"
"""
import sys
import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DB_PATH = "interns.db"
TABLE_NAME = "interns_cleaned"

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="2F5597")
TITLE_FONT = Font(bold=True, size=14, color="2F5597")

def write_table(ws, df, start_row=1):
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=start_row, column=j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    for j, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(14, len(str(col)) + 4)
    return start_row + len(df) + 1

def main(period_label: str):
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql(
        f'SELECT intern_id, intern_role, department, project_quality, '
        f'mentor_feedback, task_completion_proxy FROM {TABLE_NAME}', conn)
    conn.close()

    wb = Workbook()

    # --- Raw Data sheet (with a live Flag formula) ---
    ws_raw = wb.active
    ws_raw.title = "Raw Data"
    end_row = write_table(ws_raw, df)
    ws_raw.cell(row=1, column=7, value="Flag").font = HEADER_FONT
    ws_raw.cell(row=1, column=7).fill = HEADER_FILL
    for i in range(2, end_row):
        ws_raw.cell(row=i, column=7,
            value=f'=IF(OR(D{i}<2,E{i}<2,F{i}<2),"Needs Attention","OK")')
    last_row = end_row - 1

    # --- Summary by Intern Role ---
    ws_role = wb.create_sheet("Summary by Role")
    ws_role.cell(row=1, column=1, value=f"Monthly Intern Performance Report — {period_label}").font = TITLE_FONT
    roles = sorted(df["intern_role"].unique())
    headers = ["Intern Role", "Avg Project Quality", "Avg Mentor Feedback", "Avg Task Completion", "# Interns"]
    for j, h in enumerate(headers, start=1):
        c = ws_role.cell(row=3, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for i, role in enumerate(roles, start=4):
        ws_role.cell(row=i, column=1, value=role)
        ws_role.cell(row=i, column=2,
            value=f'=ROUND(AVERAGEIF(\'Raw Data\'!B2:B{last_row},A{i},\'Raw Data\'!D2:D{last_row}),2)')
        ws_role.cell(row=i, column=3,
            value=f'=ROUND(AVERAGEIF(\'Raw Data\'!B2:B{last_row},A{i},\'Raw Data\'!E2:E{last_row}),2)')
        ws_role.cell(row=i, column=4,
            value=f'=ROUND(AVERAGEIF(\'Raw Data\'!B2:B{last_row},A{i},\'Raw Data\'!F2:F{last_row}),2)')
        ws_role.cell(row=i, column=5,
            value=f'=COUNTIF(\'Raw Data\'!B2:B{last_row},A{i})')
    for col, w in zip("ABCDE", [28, 18, 18, 18, 12]):
        ws_role.column_dimensions[col].width = w

    # --- Summary by Department ---
    ws_dept = wb.create_sheet("Summary by Department")
    depts = sorted(df["department"].unique())
    for j, h in enumerate(["Department", "Avg Project Quality", "Avg Mentor Feedback",
                           "Avg Task Completion", "# Interns"], start=1):
        c = ws_dept.cell(row=1, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for i, dept in enumerate(depts, start=2):
        ws_dept.cell(row=i, column=1, value=dept)
        ws_dept.cell(row=i, column=2,
            value=f'=ROUND(AVERAGEIF(\'Raw Data\'!C2:C{last_row},A{i},\'Raw Data\'!D2:D{last_row}),2)')
        ws_dept.cell(row=i, column=3,
            value=f'=ROUND(AVERAGEIF(\'Raw Data\'!C2:C{last_row},A{i},\'Raw Data\'!E2:E{last_row}),2)')
        ws_dept.cell(row=i, column=4,
            value=f'=ROUND(AVERAGEIF(\'Raw Data\'!C2:C{last_row},A{i},\'Raw Data\'!F2:F{last_row}),2)')
        ws_dept.cell(row=i, column=5,
            value=f'=COUNTIF(\'Raw Data\'!C2:C{last_row},A{i})')
    for col, w in zip("ABCDE", [24, 18, 18, 18, 12]):
        ws_dept.column_dimensions[col].width = w

    out_name = f"Monthly_Intern_Report_{period_label.replace(' ', '_')}.xlsx"
    wb.save(out_name)
    print(f"Report saved: {out_name}")

if __name__ == "__main__":
    period = sys.argv[1] if len(sys.argv) > 1 else "Sample Month"
    main(period)
